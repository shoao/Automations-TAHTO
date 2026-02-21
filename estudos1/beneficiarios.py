from openpyxl import load_workbook
from openpyxl import Workbook

matriculas = []
file = load_workbook("Acessos.xlsx")
tabs = file.sheetnames

# Garante que está acessando a aba certa da planilha
if "Perda de Acesso" in tabs:
	aba_acesso = file["Perda de Acesso"]
else:
	aba_acesso = file[file.sheetnames[0]]



# Percorre até 30 matriculas e guarda os valores
for i in range(3, 30): 
	bc = aba_acesso[f"A{i}"].value

	if not bc:
		break

	else:
		matriculas.append(bc)

# remove matriculas duplicadas
matriculas = list(set(matriculas))

# cria novo arquivo de beneficiários

beneficiarios = Workbook()
aba_beneficiarios = beneficiarios.active

aba_beneficiarios["A1"].value = "MATRICULA"

for linha, matricula in enumerate(matriculas, start=2):
    aba_beneficiarios[f"A{linha}"] = matricula

beneficiarios.save("BENEFICIARIOS.xlsx")