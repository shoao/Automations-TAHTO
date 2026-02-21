from openpyxl import load_workbook
from openpyxl import Workbook

# MATRICULA NA COLUNA C

matricula = [766581, 802673, 802674]

file = load_workbook("quadroGeral 1.xlsx")
aba_gip = file.active


for linha in range(2, aba_gip.max_row + 1):
	matricula_planilha = aba_gip[f"C{linha}"].value


	if matricula_planilha in matricula:
		print("Achou!")
		regiao = aba_gip[f"A{linha}"].value
		nome = aba_gip[f"D{linha}"].value
		status = aba_gip[f"B{linha}"].value
		cargo = aba_gip[f"F{linha}"].value
		matricula_tt = aba_gip[f"AA{linha}"].value
		cod_setor = aba_gip[f"AC{linha}"].value
		nome_setor = aba_gip[f"AD{linha}"].value
		ilha_ga = aba_gip[f"CA{linha}"].value

		print(f"BC: {matricula_planilha}")
		print(f"Nome: {nome}")
		print(f"Cargo: {cargo}")
		print(f"Matricula TT: {matricula_tt}")
		print(f"Setor: {cod_setor}")
		print(f"Nome setor: {nome_setor}")
		print(f"Ilha: {ilha_ga}")
		print(f"Região: {regiao}")
		print(f"Status: {status}")



