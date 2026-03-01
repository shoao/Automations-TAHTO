from openpyxl import load_workbook
from openpyxl import Workbook

def verificarMatricula(nome_arquivo, nome_arquivo_gip):

	matriculas = []
	file = load_workbook(nome_arquivo)
	tabs = file.sheetnames

	# Garante que está acessando a aba certa da planilha
	if "Perda de Acesso" in tabs:
		aba_acesso = file["Perda de Acesso"]
	else:
		aba_acesso = file[file.sheetnames[0]]



	# Percorre até 30 matriculas e guarda os valores
	for i in range(3, 30): 
		bc = aba_acesso[f"A{i}"].value

		if not bc:
			break

		else:
			matriculas.append(bc)

	# remove matriculas duplicadas
	matriculas = list(set(matriculas))


	# PEGANDO MATRICULA ESPELHO

	aba_espelho = file[file.sheetnames[2]]

	matricula_espelho = aba_espelho["A3"].value

	# 


	print(f"Verificando matriculas: {matriculas}")
	print(f"Matricula Espelho: {matricula_espelho}\n")



	file = load_workbook(nome_arquivo_gip)
	aba_gip = file.active


	for linha in range(2, aba_gip.max_row + 1):
		matricula_planilha = aba_gip[f"C{linha}"].value


		if matricula_planilha in matriculas:
			regiao = aba_gip[f"A{linha}"].value
			nome = aba_gip[f"D{linha}"].value
			status = aba_gip[f"B{linha}"].value
			cargo = aba_gip[f"F{linha}"].value
			matricula_tt = aba_gip[f"AA{linha}"].value
			cod_setor = aba_gip[f"AC{linha}"].value
			nome_setor = aba_gip[f"AD{linha}"].value
			ilha_ga = aba_gip[f"CA{linha}"].value

			print(f"BC: {matricula_planilha}")
			print(f"Nome: {nome}")
			print(f"Cargo: {cargo}")
			print(f"Matricula TT: {matricula_tt}")
			print(f"Setor: {cod_setor}")
			print(f"Nome setor: {nome_setor}")
			print(f"Ilha: {ilha_ga}")
			print(f"Região: {regiao}")
			print(f"Status: {status}")
			print("\n")


		elif matricula_planilha == matricula_espelho:
			print("\n--- MATRICULA ESPELHO ---\n")
			regiao = aba_gip[f"A{linha}"].value
			nome = aba_gip[f"D{linha}"].value
			status = aba_gip[f"B{linha}"].value
			cargo = aba_gip[f"F{linha}"].value
			matricula_tt = aba_gip[f"AA{linha}"].value
			cod_setor = aba_gip[f"AC{linha}"].value
			nome_setor = aba_gip[f"AD{linha}"].value
			ilha_ga = aba_gip[f"CA{linha}"].value

			print(f"BC: {matricula_planilha}")
			print(f"Nome: {nome}")
			print(f"Cargo: {cargo}")
			print(f"Matricula TT: {matricula_tt}")
			print(f"Setor: {cod_setor}")
			print(f"Nome setor: {nome_setor}")
			print(f"Ilha: {ilha_ga}")
			print(f"Região: {regiao}")
			print(f"Status: {status}")
			print("\n")

	while True:
		input("Digite para sair: ")
		break
		
	return matriculas




