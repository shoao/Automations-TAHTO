from openpyxl import load_workbook
from copy import copy

def criar_aba(bairro, arquivo, estilo_cabecalho):
	if bairro not in arquivo.sheetnames:
		arquivo.create_sheet(bairro)
		nova_aba = arquivo[bairro]
		nova_aba["A1"].value = "Data de Nascimento"
		nova_aba["B1"].value = "Pessoa"
		nova_aba["C1"].value = "Bairro"

		nova_aba["A1"]._style = estilo_cabecalho
		nova_aba["B1"]._style = estilo_cabecalho
		nova_aba["C1"]._style = estilo_cabecalho

def transferir_info_aba(aba_origem, aba_destino, linha_origem):
	linha_destino = aba_destino.max_row + 1
	for coluna in range(1,4):
		celula_origem = aba_origem.cell(row=linha_origem, column=coluna)
		celula_destino = aba_destino.cell(row=linha_destino, column=coluna)
		celula_destino.value = celula_origem.value

		celula_destino._style = copy(celula_origem._style)

# Lendo o arquivo
arquivo = load_workbook("Bairros.xlsx")

# Pegando a aba que vai trabalhar
# dados = arquivo.active
print(arquivo.sheetnames)

# Pegando os dados da aba
dados = arquivo["Base de Dados"]

# Pegando a ultima linha da coluna
ultima_linha = dados.max_row
print(ultima_linha)


estilo_cabecalho = copy(dados["A1"]._style)

for linha in range(2, ultima_linha + 1):

	bairro = dados[f'C{linha}'].value
	print(bairro)

	if not bairro:
		break

	criar_aba(bairro, arquivo, estilo_cabecalho)
	aba_destino = arquivo[bairro]

	transferir_info_aba(dados, aba_destino, linha)

arquivo.save("Bairros2.xlsx")