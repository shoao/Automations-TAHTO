from openpyxl import load_workbook

arquivo = load_workbook("Alunos.xlsx")

print(arquivo.sheetnames) # função para pegar as abas da planilha

aba_atual = arquivo.active # pegar a aba que está ativa

print(aba_atual)

aba_alunos = arquivo["Planilha1"]
print(aba_alunos)

print(aba_alunos["A1"].value) # pegando valores das células

aba_alunos["A1"].value = "Alunos"

print(aba_alunos["A1"].value)

arquivo.save("Alunos2.xlsx")

# pegar ultima linha da coluna

print(aba_alunos.max_row)
print(len(aba_alunos["A"]))